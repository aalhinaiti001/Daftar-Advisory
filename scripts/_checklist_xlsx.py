"""Shared builder for the grouped-checklist workbooks shipped with articles.

Both checklist JSON files use the same shape (groups, each with ref/title/items),
so they get the same workbook: a header band, one banded row per group, and a
ref/check/owner/date/status/evidence grid underneath.

House style: cream ground, near-black header band, thin warm separators, no
vertical grid, gridlines off.
"""

import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent

INK = "FF1A1814"
RUST = "FFA8341F"
MUTED = "FF6F665D"
RULE = "FFD8D2C4"
BAND = "FFEFEBE1"

HEADERS = ["Ref", "Check", "Owner", "Target date", "Status", "Evidence / notes"]
WIDTHS = [8, 74, 18, 14, 14, 34]

DISCLAIMER = (
    "Daftar Advisory is a non-attest advisory practice, not a registered statutory auditor. "
    "This checklist is a general preparation aid, not accounting, assurance, tax or regulatory "
    "advice, and not a compliance assessment. {extra}"
)


def build(src_name: str, out_name: str, title: str, subtitle: str, sheet: str, extra: str) -> None:
    src = ROOT / "app" / "_data" / src_name
    out = ROOT / "public" / out_name

    data = json.loads(src.read_text(encoding="utf-8"))
    groups = data["groups"]
    total = sum(len(g["items"]) for g in groups)

    hairline = Border(bottom=Side(style="thin", color=RULE))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = Font(name="Inter", size=16, bold=True, color=INK)
    ws["A2"] = subtitle.format(total=total)
    ws["A2"].font = Font(name="Inter", size=10, color=MUTED)
    ws["A3"] = "Daftar Advisory · daftaradvisory.com"
    ws["A3"].font = Font(name="Inter", size=10, color=RUST)

    row = 5
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Inter", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

    for group in groups:
        row += 1
        cell = ws.cell(row=row, column=1, value=f"{group['ref']}. {group['title']}")
        cell.font = Font(name="Inter", size=11, bold=True, color=INK)
        cell.alignment = Alignment(vertical="center")
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=BAND)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        ws.row_dimensions[row].height = 22

        for i, item in enumerate(group["items"], start=1):
            row += 1
            values = [f"{group['ref']}/{i:02d}", item, "", "", "Not started", ""]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = Font(name="Inter", size=10, color=INK)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = hairline
            ws.cell(row=row, column=1).font = Font(name="JetBrains Mono", size=9, color=RUST)
            ws.row_dimensions[row].height = 30

    for c, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row += 2
    cell = ws.cell(row=row, column=1, value=DISCLAIMER.format(extra=extra))
    cell.font = Font(name="Inter", size=9, italic=True, color=MUTED)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    ws.row_dimensions[row].height = 46

    ws.freeze_panes = "A6"
    wb.save(out)
    print(f"wrote {out.relative_to(ROOT)} ({total} checks in {len(groups)} groups)")
