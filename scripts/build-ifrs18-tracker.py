#!/usr/bin/env python3
"""Builds public/ifrs-18-transition-plan-2026.xlsx, shipped with the IFRS 18 article.

Mirrors the plan table in app/knowledge/ifrs-18-transition-2026/page.tsx. Run:

    pip install openpyxl && python3 scripts/build-ifrs18-tracker.py

House style per the ahmad-house-style skill: cream ground, near-black header
band, thin warm separators, no vertical grid, gridlines off.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PAPER = "FFF4F1EA"
INK = "FF1A1814"
RUST = "FFA8341F"
MUTED = "FF6F665D"
RULE = "FFD8D2C4"

hairline = Border(bottom=Side(style="thin", color=RULE))

wb = Workbook()

# ---------------------------------------------------------------- Plan sheet
ws = wb.active
ws.title = "Transition plan"
ws.sheet_view.showGridLines = False

ws["A1"] = "IFRS 18 transition plan"
ws["A1"].font = Font(name="Inter", size=16, bold=True, color=INK)
ws["A2"] = "FY2026 comparative year · first application for periods beginning 1 January 2027"
ws["A2"].font = Font(name="Inter", size=10, color=MUTED)
ws["A3"] = "Daftar Advisory · daftaradvisory.com"
ws["A3"].font = Font(name="Inter", size=10, color=RUST)

HEADERS = [
    "#", "Workstream", "FY2026 deliverable", "Evidence source",
    "Owner", "Target date", "Status", "Notes",
]

ROWS = [
    ("01", "Specified main business activity",
     "Written determination at consolidated and separate level, with the reasoning retained",
     "Board or management paper; group structure; regulatory licences", "", "Sep 2026"),
    ("02", "Classification policy, five categories",
     "Approved policy covering operating, investing, financing, income taxes and discontinued operations",
     "Draft accounting policy; chart of accounts mapping", "", "Oct 2026"),
    ("03", "Cross-cutting judgements",
     "Documented treatment for FX, derivatives and hedges, leases, disposals and restructuring",
     "Technical memo per judgement; supporting contracts", "", "Oct 2026"),
    ("04", "By-nature disclosure feasibility",
     "Confirmation that by-nature detail exists at the required granularity, or a capture change made while 2026 is still open",
     "General ledger extract; chart of accounts; ERP configuration", "", "Oct 2026"),
    ("05", "MPM inventory",
     "Inventory built from what has actually been published, with tax and NCI effect per reconciling item",
     "Earnings releases; investor decks; management commentary; analyst materials", "", "Nov 2026"),
    ("06", "Systems and data",
     "Chart of accounts and consolidation system able to produce the new structure for the full comparative year",
     "ERP and consolidation change log; test extract", "", "Nov 2026"),
    ("07", "IAS 7 and IAS 34 knock-on",
     "Cash flow starting point and interest/dividend classification set; interim MPM reporting scoped",
     "Draft cash flow statement; interim reporting calendar", "", "Dec 2026"),
]

start = 5
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=start, column=c, value=h)
    cell.font = Font(name="Inter", size=10, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[start].height = 26

for i, (num, name, deliverable, evidence, owner, date) in enumerate(ROWS):
    r = start + 1 + i
    values = [num, name, deliverable, evidence, owner, date, "Not started", ""]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Inter", size=10, color=INK)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = hairline
    ws.cell(row=r, column=1).font = Font(name="JetBrains Mono", size=9, color=RUST)
    ws.cell(row=r, column=2).font = Font(name="Inter", size=10, bold=True, color=INK)
    ws.row_dimensions[r].height = 46

WIDTHS = [5, 30, 46, 40, 18, 14, 14, 30]
for c, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

note_row = start + len(ROWS) + 2
ws.cell(row=note_row, column=1,
        value=("Daftar Advisory is a non-attest advisory practice, not a registered statutory auditor. "
               "This tracker is a general planning aid, not accounting advice. Confirm every requirement "
               "against the issued standards and your current local regulatory instructions."))
ws.cell(row=note_row, column=1).font = Font(name="Inter", size=9, italic=True, color=MUTED)
ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
ws.row_dimensions[note_row].height = 42

ws.freeze_panes = "A6"

import pathlib
OUT = pathlib.Path(__file__).resolve().parent.parent / "public" / "ifrs-18-transition-plan-2026.xlsx"
wb.save(OUT)
print(f"wrote {OUT.name}")
