#!/usr/bin/env python3
"""Builds public/audit-readiness-checklist.xlsx from the shared checklist JSON.

The page at /knowledge/audit-readiness-checklist imports the same JSON, so the
page and the workbook cannot drift. Edit app/_data/audit-readiness-checklist.json,
then re-run:

    pip install openpyxl && python3 scripts/build-audit-checklist.py

House style: cream ground, near-black header band, thin warm separators, no
vertical grid, gridlines off.
"""

import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC = ROOT / "app" / "_data" / "audit-readiness-checklist.json"
OUT = ROOT / "public" / "audit-readiness-checklist.xlsx"

INK = "FF1A1814"
RUST = "FFA8341F"
MUTED = "FF6F665D"
RULE = "FFD8D2C4"
BAND = "FFEFEBE1"

hairline = Border(bottom=Side(style="thin", color=RULE))

data = json.loads(SRC.read_text(encoding="utf-8"))
groups = data["groups"]
total = sum(len(g["items"]) for g in groups)

wb = Workbook()
ws = wb.active
ws.title = "Audit readiness"
ws.sheet_view.showGridLines = False

ws["A1"] = "Audit readiness checklist"
ws["A1"].font = Font(name="Inter", size=16, bold=True, color=INK)
ws["A2"] = f"{total} checks across the close, schedules, judgments, controls and the request list"
ws["A2"].font = Font(name="Inter", size=10, color=MUTED)
ws["A3"] = "Daftar Advisory · daftaradvisory.com"
ws["A3"].font = Font(name="Inter", size=10, color=RUST)

HEADERS = ["Ref", "Check", "Owner", "Target date", "Status", "Evidence / notes"]
WIDTHS = [8, 74, 18, 14, 14, 34]

row = 5
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=row, column=c, value=h)
    cell.font = Font(name="Inter", size=10, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 24

for group in groups:
    row += 1
    label = f"{group['ref']}. {group['title']}"
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(name="Inter", size=11, bold=True, color=INK)
    cell.alignment = Alignment(vertical="center")
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=BAND)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    ws.row_dimensions[row].height = 22

    for i, item in enumerate(group["items"], start=1):
        row += 1
        ref = f"{group['ref']}/{i:02d}"
        for c, v in enumerate([ref, item, "", "", "Not started", ""], start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name="Inter", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = hairline
        ws.cell(row=row, column=1).font = Font(name="JetBrains Mono", size=9, color=RUST)
        ws.row_dimensions[row].height = 30

for c, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

row += 2
note = (
    "Daftar Advisory is a non-attest advisory practice, not a registered statutory auditor. "
    "This checklist is a general preparation aid, not accounting or assurance advice, and not a "
    "compliance assessment. It does not replace the requirements of the applicable auditing "
    "standards. Confirm what your audit requires against the issued standards and your auditor's "
    "own request list."
)
cell = ws.cell(row=row, column=1, value=note)
cell.font = Font(name="Inter", size=9, italic=True, color=MUTED)
cell.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
ws.row_dimensions[row].height = 46

ws.freeze_panes = "A6"

wb.save(OUT)
print(f"wrote {OUT.relative_to(ROOT)} ({total} checks in {len(groups)} groups)")
